#!/usr/bin/env python3

import logging
import datetime
from urllib import parse
from selenium.webdriver.support.wait import WebDriverWait
from api.core import SeleniumEtl, Schedule
from api.database import etl_central_session
from api.environment import EDGE_BASE_URL
from api.uhl_etl.edge import login
from lbrc_edge import EdgeSiteStudy
from time import sleep
import csv
import tempfile
from openpyxl import load_workbook


class EdgeSiteStudyDownload(SeleniumEtl):

    REPORT_FOLDER = 'ProjectAttributeReport'
    REPORT = 'BRC Report (Richard)'

    def __init__(self):
        super().__init__(schedule=Schedule.daily_at_4am)

    def do_selenium_etl(self, driver):
        self._studies = self.get_studies(driver)

    def do_post_selenium_etl(self):
        with etl_central_session() as session:
            self.log("Deleting old studies")

            session.query(EdgeSiteStudy).delete()

            self.log("Creating new studies")
            session.add_all(self._studies)

    def get_studies(self, driver):
        self.log("Getting studies")

        result = []

        login(driver)

        driver.get(parse.urljoin(
            EDGE_BASE_URL,
            self.REPORT_FOLDER,
        ))

        driver.find_element_by_css_selector('input[value="Load query"]').click()
        sleep(5)
        driver.find_element_by_xpath(f'//a[@name="linkLoadQuery" and text()="{self.REPORT}"]').click()
        sleep(5)
        driver.find_element_by_css_selector('input[value="Submit query"]').click()
        sleep(15)
        driver.find_element_by_id('butDownloadCsv').click()
        sleep(30)

        download_file = tempfile.NamedTemporaryFile()

        try:
            # list all the completed remote files (waits for at least one)
            files = WebDriverWait(driver, 30, 1).until(lambda driver: self.get_downloaded_files(driver))

            # get the content of the first file remotely
            content = self.get_file_content(driver, files[0])

            with open(download_file.name, 'wb') as f:
                f.write(content)            

        finally:
            logging.info("Test test_grid_download_files executed successfully")

        with open(download_file.name, encoding="utf-16-le") as csvfile:
            study_details = csv.DictReader(csvfile, delimiter=',', quotechar='"')

            for row in study_details:

                if row.get('Primary Clinical Management Areas', '').upper() not in ['CARDIOLOGY', 'VASCULAR SERVICES', 'CARDIAC SURGERY']:
                    continue

                e = EdgeSiteStudy(
                    project_id=self.int_or_none(row['Project ID']),
                    iras_number=self.string_or_none(row['IRAS Number']),
                    project_short_title=self.string_or_none(row['Project Short title']),
                    primary_clinical_management_areas=self.string_or_none(row['Primary Clinical Management Areas']),
                    project_site_status=self.string_or_none(row['Project site status']),
                    project_site_rand_submission_date=self.date_or_none(row['Project site R&D Submission Date']),
                    project_site_start_date_nhs_permission=self.date_or_none(row['Project site Start date (NHS Permission)']),
                    project_site_date_site_confirmed=self.date_or_none(row['Project site date site confirmed']),
                    project_site_planned_closing_date=self.date_or_none(row['Project site Planned closing date']),
                    project_site_closed_date=self.date_or_none(row['Project site Closed date']),
                    project_site_planned_recruitment_end_date=self.date_or_none(row['Project site planned recruitment end date']),
                    project_site_actual_recruitment_end_date=self.date_or_none(row['Project site actual recruitment end date']),
                    principal_investigator=self.name_or_none(row['Principal Investigator']),
                    project_site_target_participants=self.int_or_none(row['Project site target participants']),
                    recruited_org=self.int_or_none(row['Recruited (org)']),
                    project_site_lead_nurses=self.name_or_none(row['Project site lead nurse(s)']),
                )

                e.calculate_values()

                result.append(e)

        self.log("Getting studies: COMPLETED")

        download_file.close()

        return result

    def string_or_none(self, string_element):
        string_element = string_element.strip()
        if string_element:
            return string_element
        else:
            return None

    def name_or_none(self, string_element):
        string_element = string_element.strip()
        if string_element:
            name = ' '.join(reversed(
                [p.strip() for p in filter(lambda x: len(x) > 0, string_element.split(','))]
            )).strip()

            if name:
                return name

    def int_or_none(self, int_element):
        int_string = int_element.strip()
        if int_string:
            return int(int_string)
        else:
            return None

    def date_or_none(self, date_element):
        date_string = date_element.strip()
        if date_string:
            return datetime.datetime.strptime(date_string, "%d/%m/%Y").date()
        else:
            return None

    def boolean_or_none(self, boolean_element):
        boolean_string = boolean_element.strip().upper()
        if boolean_string in ['YES', 'TRUE', '1']:
            return True
        elif boolean_string in ['NO', 'FALSE', '0']:
            return False
        else:
            return None


class EdgeStudyOverviewDownload(SeleniumEtl):

    REPORT_FOLDER = 'ProjectOverviewReport'

    def __init__(self):
        super().__init__(schedule=Schedule.daily_7pm)

    def do_selenium_etl(self, driver):
        self._studies = self.get_studies(driver)

    def do_post_selenium_etl(self):
        pass
        # with etl_central_session() as session:
        #     self.log("Deleting old studies")

        #     session.query(EdgeSiteStudy).delete()

        #     self.log("Creating new studies")
        #     session.add_all(self._studies)

    def get_studies(self, driver):
        self.log("Getting studies")

        result = []

        login(driver)

        driver.get(parse.urljoin(
            EDGE_BASE_URL,
            self.REPORT_FOLDER,
        ))

        driver.find_element_by_xpath(f'//a[normalize-space(text())="Download Excel"]').click()
        sleep(30)

        download_file = tempfile.NamedTemporaryFile(suffix='.xlsx')

        try:
            # list all the completed remote files (waits for at least one)
            files = WebDriverWait(driver, 30, 1).until(lambda driver: self.get_downloaded_files(driver))

            # get the content of the first file remotely
            content = self.get_file_content(driver, files[0])

            with open(download_file.name, 'wb') as f:
                f.write(content)            

            # with open('test.xlsx', 'wb') as f:
            #     f.write(content)            

        finally:
            logging.info("Test test_grid_download_files executed successfully")

        wb = load_workbook(filename=download_file.name, read_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)

        headers = next(rows)

        for r in rows:
            dict(zip(headers, r))

        # with open(download_file.name, encoding="utf-16-le") as csvfile:
        #     study_details = csv.DictReader(csvfile, delimiter=',', quotechar='"')

        #     for row in study_details:

        #         if row.get('Primary Clinical Management Areas', '').upper() not in ['CARDIOLOGY', 'VASCULAR SERVICES', 'CARDIAC SURGERY']:
        #             continue

        #         e = EdgeSiteStudy(
        #             project_id=self.int_or_none(row['Project ID']),
        #             iras_number=self.string_or_none(row['IRAS Number']),
        #             project_short_title=self.string_or_none(row['Project Short title']),
        #             primary_clinical_management_areas=self.string_or_none(row['Primary Clinical Management Areas']),
        #             project_site_status=self.string_or_none(row['Project site status']),
        #             project_site_rand_submission_date=self.date_or_none(row['Project site R&D Submission Date']),
        #             project_site_start_date_nhs_permission=self.date_or_none(row['Project site Start date (NHS Permission)']),
        #             project_site_date_site_confirmed=self.date_or_none(row['Project site date site confirmed']),
        #             project_site_planned_closing_date=self.date_or_none(row['Project site Planned closing date']),
        #             project_site_closed_date=self.date_or_none(row['Project site Closed date']),
        #             project_site_planned_recruitment_end_date=self.date_or_none(row['Project site planned recruitment end date']),
        #             project_site_actual_recruitment_end_date=self.date_or_none(row['Project site actual recruitment end date']),
        #             principal_investigator=self.name_or_none(row['Principal Investigator']),
        #             project_site_target_participants=self.int_or_none(row['Project site target participants']),
        #             recruited_org=self.int_or_none(row['Recruited (org)']),
        #             project_site_lead_nurses=self.name_or_none(row['Project site lead nurse(s)']),
        #         )

        #         result.append(e)

        self.log("Getting studies: COMPLETED")

        download_file.close()

        return result

    def string_or_none(self, string_element):
        string_element = string_element.strip()
        if string_element:
            return string_element
        else:
            return None

    def name_or_none(self, string_element):
        string_element = string_element.strip()
        if string_element:
            name = ' '.join(reversed(
                [p.strip() for p in filter(lambda x: len(x) > 0, string_element.split(','))]
            )).strip()

            if name:
                return name

    def int_or_none(self, int_element):
        int_string = int_element.strip()
        if int_string:
            return int(int_string)
        else:
            return None

    def date_or_none(self, date_element):
        date_string = date_element.strip()
        if date_string:
            return datetime.datetime.strptime(date_string, "%d/%m/%Y").date()
        else:
            return None

    def boolean_or_none(self, boolean_element):
        boolean_string = boolean_element.strip().upper()
        if boolean_string in ['YES', 'TRUE', '1']:
            return True
        elif boolean_string in ['NO', 'FALSE', '0']:
            return False
        else:
            return None
